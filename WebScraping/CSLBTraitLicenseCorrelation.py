import requests
from bs4 import BeautifulSoup
import pandas as pd
import re
import openpyxl

url = "https://www.cslb.ca.gov/about_us/library/licensing_classifications/"
response = requests.get(url)

soup = BeautifulSoup(response.content, 'html.parser')

links = soup.find_all('a', href=lambda href: href and "Licensing_Classifications_Detail.aspx?Class=" in href)

hrefs = [link['href'] for link in links]

licenseTexts = []

for license in hrefs[4:]:
  licensePage = url + license
  linkResponse = requests.get(licensePage)
  souping = BeautifulSoup(linkResponse.content, 'html.parser')
  section = souping.find("div", class_="main-primary")
  paragraphs = section.find_all("p")
  text = [paragraph.get_text() for paragraph in paragraphs]
  licenseTexts.append(text)

cleanedLicenseTexts = []

for licenseText in licenseTexts:
  cleansing = []
  for paragraph in licenseText:
    cleanedString = paragraph.replace("\n", "").replace("\t", "").replace("\r", "").replace("            "," ")
    cleansing.append(cleanedString)
  cleanedLicenseTexts.append(cleansing)

soup2 = BeautifulSoup(response.content, 'html.parser')

linksSpecialty = soup2.find('a',href="C-61_Limited_Speciality/Default.aspx")

specialties = url + linksSpecialty['href']

response2 = requests.get(specialties)
souping2 = BeautifulSoup(response2.content, 'html.parser')

# Method 1: All hrefs in Page
hrefs2 = souping2.find_all('a', href=lambda href: href and "Licensing_Classifications_Detail.aspx?Class=" in href)

# Method 2: Specific locations
ulClass = souping2.find('ul',class_='list-understated')

listElements = ulClass.find_all('li')

linksforEach = []

for element in listElements:
  filter = []
  for finiteElement in element.find_all('a'):   
    if finiteElement['href'][0:2] == "..":
      filter.append(finiteElement['href'][2:])
    else:
      filter.append(finiteElement['href'])
  linksforEach.append(filter)

# Text Titles only
textlistElements = [element.get_text() for element in listElements]

specialtiesTitlesDescriptions = []

for x, y in zip(linksforEach, textlistElements):
  transportingStrings = []
  transportingStrings.append(y)
  for j in range(len(x)):
    direction = url + x[j]
    response3 = requests.get(direction)
    souping3 = BeautifulSoup(response3.content, 'html.parser')
    classing = souping3.find("div", class_="main-primary")
    try:
      paragraphs = classing.find_all("p")
    except AttributeError:
      continue
    textSpecialty = [paragraph.get_text() for paragraph in paragraphs]
    cleansing2 = []
    for paragraph in textSpecialty:
      cleanedString = paragraph.replace("\n", "").replace("\t", "").replace("\r", "")
      cleanFurther = " ".join(cleanedString.split())
      cleansing2.append(cleanFurther)
    transportingStrings.append(cleansing2)
  specialtiesTitlesDescriptions.append(transportingStrings)

specialtiesExcelTransition = []

for specialtyCleaned in specialtiesTitlesDescriptions:
  agentTransport = []
  for specialtyAttribute in specialtyCleaned:
    if type(specialtyAttribute) == list:
      for attributeItself in specialtyAttribute:
        agentTransport.append(attributeItself)
    else:
      cleanedString = specialtyAttribute.replace("-", "")
      anotherClean = cleanedString.split("  ")
      agentTransport.append(anotherClean[0])
      try: 
        agentTransport.append(anotherClean[1])
      except IndexError:
        continue
  specialtiesExcelTransition.append(agentTransport)

excelFile = "/Users/damiamalfaro/Desktop/Sonder/Code/MiscellaneousAutomations/CSLBTraitLicenseCorrelation/CSLBTraitLicensCorrelation.xlsx"
workbook = openpyxl.load_workbook(excelFile)
sheet_name = "Sheet1"
sheet=workbook[sheet_name]
row_start = sheet.max_row+1
for row_index, row_data in enumerate(specialtiesExcelTransition, start=row_start):
    for col_index, value in enumerate(row_data, start=1):
        sheet.cell(row=row_index, column=col_index, value=value)

workbook.save(excelFile)
print("Atta boy")
