import PyPDF2
import pdfplumber
from pytesseract import image_to_string
import openpyxl


class Specs():

    
    # Function 1.1
    def SectionsPDFText(self):

        raw_text = []

        with open(self,"rb") as file:
            read = PyPDF2.PdfReader(file)

            pages = len(read.pages)
            print(f"{pages} pages")

            for page in range(pages):
                page_text = read.pages[page]
                raw_text.append(page_text.extract_text())


        cleanup = [dividors.split("\n") for dividors in raw_text]
        sections = []

        for x in cleanup:
            for y in x:
                if y[:1].isdigit() == True and len(y) > 2:
                    sections.append(y)

                else:
                    continue

        return sections

    # Function 1.2: The reason why this is different is because each pdf varies, therefore each cleanup varies 
    def CleaningExtraChar(self):
        crust = [spec.replace(".","")[:-6] for spec in self]
        mantle = [magma.replace(" ","") for magma in crust]
        asthenosphere = [lava.replace("-","") for lava in mantle]

        return asthenosphere
    
    # Function 1.3: Renaming the Tabs in the Excel Spreadsheet based on Specs Text
    def RenameTabs(self,specs):
        path = self
        workbook = openpyxl.load_workbook(path)
        spec_names = specs
        index_halt = 8
        if len(workbook.sheetnames) < index_halt + len(spec_names):
            print("Add more Templates")
        else:
            for i, name in enumerate(spec_names):
                current_tab_index = index_halt + i
                workbook[workbook.sheetnames[current_tab_index]].title = name


            workbook.save(self)

        workbook.close()

    
    # Function 1.4: Allocates list of names to a tab for simplicity 
    def AllocatingSpecsNames(self,specs):
        workbook = openpyxl.load_workbook(self)
        sheet = workbook["BidLayout"]

        for index, value in enumerate(specs,start=1):
            cell = f"A{index}"
            sheet[cell] = value

        workbook.save(self)




    def SectionsOCR(self):
        with pdfplumber.open(self) as pdf:
            for page in pdf.pages:
                page_image = page.to_image()
                text = image_to_string(page_image.original)
                print(text)



if __name__ == "__main__":
    # Insert corresponding Specs PDF File
    specs_file = "/Users/damiamalfaro/Desktop/Accuracy/Estimate/DataBase/GoletaTraintDepot/GoletaSpecSections.pdf"
    
    # Insert corresponding Excel Spreadsheet
    excel_file = "/Users/damiamalfaro/Desktop/Accuracy/Estimate/DataBase/GoletaTraintDepot/GoletaTrainDepot-Spreadsheet.xlsm"
   
    # Function 1: Reads the specs and acquires a list of all spec sections if file is PDF Text-based.
    RawSpecs = Specs.CleaningExtraChar(Specs.SectionsPDFText(specs_file))

    # Function 1.3: Renaming the Tabs
    #Specs.RenameTabs(excel_file,RawSpecs)

    # Function 1.4: Allocating Specs Names on a Column for VBA
    Specs.AllocatingSpecsNames(excel_file,RawSpecs)

    # Function 2: Reads the specs and acquires a list of all spec sections if file is Image-based.
    #Specs.SectionsOCR(specs_file)
    













